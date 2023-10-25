from openpyxl import Workbook,load_workbook
wb=load_workbook('data_files/EmployeeData.xlsl')
ws=wb.active
max_row = ws.max_row
n = 0

for row in range(2,max_row + 1):
    gender = ws['D'+str(row)].value
    rate = ws['B'+str(row)].value
    hours = ws['C'+str(row)].value
    
    if(gender == "male"):
        #print("People id " + str(id) + " His salary is " + str(salary))
        n=n+1
print(n)
wb.close() 
